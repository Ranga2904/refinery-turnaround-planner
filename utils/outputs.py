"""
Output generation utilities for turnaround planning.

Generates:
- Excel workbooks (risk analysis, optimized scope, deferred items)
- PDF executive memos
- Summary dashboards
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict
import io


def style_header_row(ws, row_num=1):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_borders(ws, start_row, end_row, start_col, end_col):
    """Add borders to a range of cells"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def format_currency_column(ws, col_letter, start_row, end_row):
    """Format column as currency"""
    for row in range(start_row, end_row + 1):
        cell = ws[f'{col_letter}{row}']
        cell.number_format = '$#,##0'


def create_excel_workbook(
    risk_analysis: pd.DataFrame,
    optimized_scope: pd.DataFrame,
    deferred_items: pd.DataFrame,
    optimization_summary: Dict,
    constraints: Dict,
    session_id: str
) -> Workbook:
    """
    Create comprehensive Excel workbook with all analysis results.
    
    Returns:
        openpyxl Workbook object
    """
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Executive Summary
    ws_summary = wb.create_sheet("Executive Summary", 0)
    _create_summary_sheet(ws_summary, optimization_summary, constraints, session_id)
    
    # Sheet 2: Risk-Ranked Scope
    ws_risk = wb.create_sheet("Risk Analysis", 1)
    _create_risk_sheet(ws_risk, risk_analysis)
    
    # Sheet 3: Included Scope
    ws_included = wb.create_sheet("Included Scope", 2)
    _create_included_sheet(ws_included, optimized_scope)
    
    # Sheet 4: Deferred Items
    if len(deferred_items) > 0:
        ws_deferred = wb.create_sheet("Deferred Items", 3)
        _create_deferred_sheet(ws_deferred, deferred_items)
    
    return wb


def _create_summary_sheet(ws, summary: Dict, constraints: Dict, session_id: str):
    """Create executive summary sheet"""
    
    # Title
    ws['A1'] = "TURNAROUND SCOPE OPTIMIZATION SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Session info
    ws['A3'] = "Session ID:"
    ws['B3'] = session_id
    ws['A4'] = "Generated:"
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # Constraints
    ws['A6'] = "CONSTRAINTS"
    ws['A6'].font = Font(bold=True, size=12)
    ws['A7'] = "Budget Cap:"
    ws['B7'] = constraints['budget_cap']
    ws['B7'].number_format = '$#,##0'
    ws['A8'] = "Duration Cap:"
    ws['B8'] = f"{constraints['duration_cap']} days"
    ws['A9'] = "Production Capacity:"
    ws['B9'] = f"{constraints['production_capacity_bbl_day']:,.0f} bbl/day"
    ws['A10'] = "Margin:"
    ws['B10'] = f"${constraints['margin_per_bbl']}/bbl"
    
    # Results
    ws['A12'] = "OPTIMIZATION RESULTS"
    ws['A12'].font = Font(bold=True, size=12)
    
    results_data = [
        ["Items Included:", summary['total_included']],
        ["Items Deferred:", summary['total_deferred']],
        ["Total Cost:", summary['total_cost'], '$#,##0'],
        ["Budget Utilization:", f"{summary['budget_utilization_pct']:.1f}%"],
        ["Estimated Duration:", f"{summary['total_duration_parallel']:.1f} days"],
        ["Risk Avoided:", summary['total_risk_avoided'], '$#,##0'],
        ["Residual Risk:", summary['residual_risk'], '$#,##0']
    ]
    
    for idx, row in enumerate(results_data, start=13):
        ws[f'A{idx}'] = row[0]
        ws[f'B{idx}'] = row[1]
        if len(row) > 2:
            ws[f'B{idx}'].number_format = row[2]
    
    # ROI calculation
    ws['A21'] = "RETURN ON INVESTMENT"
    ws['A21'].font = Font(bold=True, size=12)
    ws['A22'] = "Risk Avoided per $1K Spent:"
    if summary['total_cost'] > 0:
        roi = summary['total_risk_avoided'] / (summary['total_cost'] / 1000)
        ws['B22'] = f"${roi:,.2f}"
    else:
        ws['B22'] = "N/A"
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def _create_risk_sheet(ws, risk_df: pd.DataFrame):
    """Create risk analysis sheet"""
    
    # Title
    ws['A1'] = "RISK ANALYSIS - ALL SCOPE ITEMS"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:J1')
    
    # Select and reorder columns
    columns_to_display = [
        'rank', 'equipment_id', 'work_type', 'cost_estimate_usd',
        'duration_days', 'failure_probability', 'consequence_usd',
        'expected_loss_avoided', 'risk_tier', 'confidence'
    ]
    
    display_df = risk_df[columns_to_display].copy()
    
    # Rename for readability
    display_df.columns = [
        'Rank', 'Equipment ID', 'Work Type', 'Cost ($)',
        'Duration (days)', 'Failure Prob', 'Consequence ($)',
        'Expected Loss ($)', 'Risk Tier', 'Confidence'
    ]
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Style header
    style_header_row(ws, row_num=3)
    
    # Format currency columns
    format_currency_column(ws, 'D', 4, len(display_df) + 3)
    format_currency_column(ws, 'G', 4, len(display_df) + 3)
    format_currency_column(ws, 'H', 4, len(display_df) + 3)
    
    # Add borders
    add_borders(ws, 3, len(display_df) + 3, 1, len(columns_to_display))
    
    # Auto-width columns
    # Auto-width columns
    for col_idx in range(1, len(columns_to_display) + 1):
        max_length = 0
        column_letter = ws.cell(row=3, column=col_idx).column_letter
        for row in range(3, len(display_df) + 4):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 20)


def _create_included_sheet(ws, included_df: pd.DataFrame):
    """Create included scope sheet"""
    
    ws['A1'] = "INCLUDED SCOPE ITEMS"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:H1')
    
    columns_to_display = [
        'rank', 'equipment_id', 'work_type', 'description',
        'cost_estimate_usd', 'duration_days', 'expected_loss_avoided', 'confidence'
    ]
    
    display_df = included_df[columns_to_display].copy()
    display_df.columns = [
        'Rank', 'Equipment ID', 'Work Type', 'Description',
        'Cost ($)', 'Duration (days)', 'Risk Avoided ($)', 'Confidence'
    ]
    
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    style_header_row(ws, row_num=3)
    format_currency_column(ws, 'E', 4, len(display_df) + 3)
    format_currency_column(ws, 'G', 4, len(display_df) + 3)
    add_borders(ws, 3, len(display_df) + 3, 1, len(columns_to_display))
    
    # Auto-width columns
    for col_idx in range(1, len(columns_to_display) + 1):
        max_length = 0
        column_letter = ws.cell(row=3, column=col_idx).column_letter
        for row in range(3, len(display_df) + 4):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 20)


def _create_deferred_sheet(ws, deferred_df: pd.DataFrame):
    """Create deferred items sheet"""
    
    ws['A1'] = "DEFERRED SCOPE ITEMS"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:G1')
    
    columns_to_display = [
        'rank', 'equipment_id', 'work_type', 'cost_estimate_usd',
        'expected_loss_avoided', 'deferral_reason', 'confidence'
    ]
    
    display_df = deferred_df[columns_to_display].copy()
    display_df.columns = [
        'Rank', 'Equipment ID', 'Work Type', 'Cost ($)',
        'Residual Risk ($)', 'Deferral Reason', 'Confidence'
    ]
    
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    style_header_row(ws, row_num=3)
    format_currency_column(ws, 'D', 4, len(display_df) + 3)
    format_currency_column(ws, 'E', 4, len(display_df) + 3)
    add_borders(ws, 3, len(display_df) + 3, 1, len(columns_to_display))
    
    # Auto-width columns
    for col_idx in range(1, len(columns_to_display) + 1):
        max_length = 0
        column_letter = ws.cell(row=3, column=col_idx).column_letter
        for row in range(3, len(display_df) + 4):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 20)


def save_excel_workbook(
    wb: Workbook,
    filename: str = None,
    session_id: str = None
) -> str:
    """
    Save workbook to file.
    
    Returns:
        Path to saved file
    """
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_part = f"_{session_id}" if session_id else ""
        filename = f"data/turnaround_analysis{session_part}_{timestamp}.xlsx"
    
    wb.save(filename)
    return filename

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER


def create_pdf_memo(
    memo_text: str,
    optimization_summary: Dict,
    constraints: Dict,
    session_id: str,
    filename: str = None
) -> str:
    """
    Create PDF executive memo.
    
    Args:
        memo_text: Markdown-formatted memo from Agent 3
        optimization_summary: Summary statistics
        constraints: Budget/duration constraints
        session_id: Session identifier
        filename: Output filename (optional)
    
    Returns:
        Path to saved PDF
    """
    
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"data/executive_memo_{session_id}_{timestamp}.pdf"
    
    # Create PDF document
    doc = SimpleDocTemplate(filename, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=20
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Header
    story.append(Paragraph("TURNAROUND SCOPE RECOMMENDATION", title_style))
    story.append(Spacer(1, 0.2 * inch))
    
    # Session info
    info_data = [
        ['Session ID:', session_id],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['Budget Cap:', f"${constraints['budget_cap']:,.0f}"],
        ['Duration Cap:', f"{constraints['duration_cap']} days"]
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 0.3 * inch))
    
    # Key metrics box
    metrics_data = [
        ['Items Included', 'Total Cost', 'Risk Avoided'],
        [
            str(optimization_summary['total_included']),
            f"${optimization_summary['total_cost']:,.0f}",
            f"${optimization_summary['total_risk_avoided']:,.0f}"
        ]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, 1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    story.append(metrics_table)
    story.append(Spacer(1, 0.3 * inch))
    
    # Parse memo text (simple markdown to PDF)
    lines = memo_text.split('\n')
    
    for line in lines:
        line = line.strip()
        
        if not line:
            story.append(Spacer(1, 0.1 * inch))
        elif line.startswith('# '):
            # H1 heading
            story.append(Paragraph(line[2:], title_style))
        elif line.startswith('## '):
            # H2 heading
            story.append(Paragraph(line[3:], heading_style))
        elif line.startswith('- ') or line.startswith('* '):
            # Bullet point
            story.append(Paragraph(f"• {line[2:]}", styles['Normal']))
        elif line.startswith('1. ') or line.startswith('2. ') or line.startswith('3. '):
            # Numbered list
            story.append(Paragraph(line, styles['Normal']))
        else:
            # Regular paragraph
            story.append(Paragraph(line, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    
    return filename


# Update the test section at the bottom:
if __name__ == "__main__":
    """Test Excel and PDF generation"""
    import pandas as pd
    
    print("=" * 60)
    print("TESTING OUTPUT GENERATION")
    print("=" * 60)
    
    # Load data
    risk_analysis = pd.read_csv('data/agent1_risk_analysis.csv')
    optimized_scope = pd.read_csv('data/agent2_included_scope.csv')
    
    # Handle empty deferred items
    try:
        deferred_items = pd.read_csv('data/agent2_deferred_items.csv')
    except:
        deferred_items = pd.DataFrame()
    
    optimization_summary = {
        'total_included': len(optimized_scope),
        'total_deferred': len(deferred_items),
        'total_cost': optimized_scope['cost_estimate_usd'].sum(),
        'total_duration_parallel': optimized_scope['duration_days'].sum() * 0.5,
        'total_risk_avoided': optimized_scope['expected_loss_avoided'].sum(),
        'residual_risk': deferred_items['expected_loss_avoided'].sum() if len(deferred_items) > 0 else 0,
        'budget_utilization_pct': (optimized_scope['cost_estimate_usd'].sum() / 18_500_000) * 100
    }
    
    constraints = {
        'budget_cap': 18_500_000,
        'duration_cap': 42,
        'production_capacity_bbl_day': 100000,
        'margin_per_bbl': 25.0
    }
    
    # Test Excel
    print("\n[1] Creating Excel workbook...")
    wb = create_excel_workbook(
        risk_analysis,
        optimized_scope,
        deferred_items,
        optimization_summary,
        constraints,
        "test-session-001"
    )
    
    excel_file = save_excel_workbook(wb, session_id="test")
    print(f"✓ Excel: {excel_file}")
    
    # Test PDF
    print("\n[2] Creating PDF memo...")
    
    # Load or create sample memo
    try:
        with open('data/agent3_executive_memo.md', 'r') as f:
            memo_text = f.read()
    except:
        memo_text = """# Turnaround Scope Recommendation

## Recommendation
Proposed scope addresses critical equipment needs within budget constraints.

## Key Inclusions
- Equipment A: High priority
- Equipment B: Medium priority

## Confidence
Analysis based on historical data with high confidence."""
    
    pdf_file = create_pdf_memo(
        memo_text,
        optimization_summary,
        constraints,
        "test-session-001"
    )
    print(f"✓ PDF: {pdf_file}")
    
    print("\n" + "=" * 60)
    print("✓ OUTPUT GENERATION COMPLETE")
    print("=" * 60)
    

if __name__ == "__main__":
    """Test Excel generation"""
    import pandas as pd
    
    print("=" * 60)
    print("TESTING EXCEL EXPORT")
    print("=" * 60)
    
    # Load data
    risk_analysis = pd.read_csv('data/agent1_risk_analysis.csv')
    optimized_scope = pd.read_csv('data/agent2_included_scope.csv')
    
    # Handle empty deferred items
    try:
        deferred_items = pd.read_csv('data/agent2_deferred_items.csv')
    except:
        deferred_items = pd.DataFrame()
    
    optimization_summary = {
        'total_included': len(optimized_scope),
        'total_deferred': len(deferred_items),
        'total_cost': optimized_scope['cost_estimate_usd'].sum(),
        'total_duration_parallel': optimized_scope['duration_days'].sum() * 0.5,
        'total_risk_avoided': optimized_scope['expected_loss_avoided'].sum(),
        'residual_risk': deferred_items['expected_loss_avoided'].sum() if len(deferred_items) > 0 else 0,
        'budget_utilization_pct': (optimized_scope['cost_estimate_usd'].sum() / 18_500_000) * 100
    }
    
    constraints = {
        'budget_cap': 18_500_000,
        'duration_cap': 42,
        'production_capacity_bbl_day': 100000,
        'margin_per_bbl': 25.0
    }
    
    # Create workbook
    print("\n[1] Creating Excel workbook...")
    wb = create_excel_workbook(
        risk_analysis,
        optimized_scope,
        deferred_items,
        optimization_summary,
        constraints,
        "test-session-001"
    )
    
    # Save
    print("[2] Saving workbook...")
    filename = save_excel_workbook(wb, session_id="test")
    
    print(f"\n✓ Excel workbook saved: {filename}")
    print(f"  Sheets: {wb.sheetnames}") 
